import argparse
from time import sleep
import openpyxl

from StartGGClient import StartGGClient


if __name__ == '__main__':
    parser = argparse.ArgumentParser(prog='AllEntrants.py')
    parser.add_argument('--token', required=True,
                        help='Start GG API token')
    parser.add_argument('--input', required=True,
                        help='input file listing tournament urls')
    
    args = parser.parse_args()

    start_client = StartGGClient(args.token)

    dataframe = openpyxl.load_workbook(args.input)
    dataframe1 = dataframe.active

    participants = {}
    tags = {}

    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            return_dict = start_client.query_participants(col[row].value.rsplit(sep='/',maxsplit=1)[-1])
            for participant in return_dict:
                if not participant['gamerTag'] in participants:
                    participants[participant['gamerTag']] = 1
                else:
                    participants[participant['gamerTag']] += 1

                if participant['prefix'] == None or participant['prefix'] == '':
                    continue
                elif not participant['prefix'] in tags:
                    tags[participant['prefix']] = 1
                else:
                    tags[participant['prefix']] += 1
            sleep(.75)
    
    participants = sorted(participants.items(), key=lambda item: item[1], reverse=True)
    f = open('participants.txt', 'w', encoding="utf-8")
    for entry in participants:
        f.write(entry[0] + ' ' + str(entry[1])+'\n')
    f.close()

    print('\n')

    f = open('tags.txt', 'w', encoding="utf-8")
    tags = sorted(tags.items(), key=lambda item: item[1], reverse=True)
    for entry in tags:
        f.write(entry[0] + ' ' + str(entry[1])+'\n')
    f.close()